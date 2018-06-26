# -*- coding: utf-8 -*-
import os
import stat
import xlwt
from xlwt import *
import openpyxl
from collections import Iterable


class ExcelHandle(object):
    # def read_excel(self, fname):
    #     book = xlrd.open_workbook(fname)
    #     sheets = book.sheet_names()                     # 获取所有表名
    #     sheet = book.sheet_by_name(sheets[0])
    #     nrows = sheet.nrows                             # sheet.nrows: 行数
    #     ncols = sheet.ncols                             # sheet.ncols: 列数
    #     for i in range(0, nrows):
    #         # row = sheet.row(i)
    #         # rowValue = sheet.row_values(i)
    #         for j in range(0, ncols):
    #             print(sheet.cell_value(i, j), '\t', end='')
    #         print()

    @staticmethod
    def full_path(fpath, fname):
        full_path = os.path.join(fpath, fname)
        if not os.path.exists(fpath):
            os.makedirs(fpath)
            os.chmod(fpath, stat.S_IRWXG)
        return full_path

    @staticmethod
    def column_map(index):
        # from collections import ChainMap
        # aa = {'0':'A', '1':'B', '2':'C', '3':'D', '4':'E', '5':'F', '6':'G', \
        #     '7':'H', '8':'I', '9':'J', '10':'K', '11':'L', '12':'M', '13':'N'}
        # c = ChainMap(aa)
        C = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N')
        return C[index]

    @staticmethod
    def to_unicode(value):
        if not value:
            return value
        elif isinstance(value, unicode):
            return value
        elif isinstance(value, str):
            result = ''
            try:
                result = value.decode('utf-8')
            except Exception as e:
                try:
                    result = value.decode('gbk')
                except Exception as e:
                    print(e)
            return result
        else:
            return value

    @staticmethod
    def add_style(sheet, cols, set_cols=None, sflag=False):
        if not set_cols:
            set_cols = [u'地址', u'资产名称', u'添加/发现时间', u'管理员', u'安全状态']

        if sflag == 1 or sflag == '1':
            style = xlwt.XFStyle()  # 创建样式
            alignment = xlwt.Alignment()  #创建居中
            alignment.horz = xlwt.Alignment.HORZ_CENTER  # 可取值: HORZ_GENERAL, HORZ_LEFT, HORZ_CENTER, HORZ_RIGHT, HORZ_FILLED, HORZ_JUSTIFIED, HORZ_CENTER_ACROSS_SEL, HORZ_DISTRIBUTED
            alignment.vert = xlwt.Alignment.VERT_CENTER  # 可取值: VERT_TOP, VERT_CENTER, VERT_BOTTOM, VERT_JUSTIFIED, VERT_DISTRIBUTED
            style.alignment = alignment  # 给样式添加文字居中属性
            # style.font.height = 430  # 设置字体大小
        
            for j, col_value in enumerate(cols):
                for col in set_cols:
                    if col == col_value:
                        sheet.col(j).width = 4000  # 设置第j列的宽
            return style
        elif sflag == 2 or sflag == '2':
            for j, col_value in enumerate(cols):
                for col in set_cols:
                    if col == col_value:
                        C = ExcelHandle.column_map(j)
                        sheet.column_dimensions[C].width = 13

    @staticmethod
    def get_default(value, v_type):
        if not value:
            v_type = v_type.lower()
            if v_type.startswith('h'):
                default = (u'序号', u'地址', u'资产名称', u'添加/发现时间', u'管理员', u'安全状态', u'可用性', u'备注')
            elif v_type.startswith('b'):
                default = ('id', 'address', 'name', 'found_time', 'administrator', 'safe_state', 'availability', 'description')
            else:
                default = tuple()
        else:
            default = tuple(value)
        return default

    # 2007版以前的Excel（xls结尾的），需要使用xlrd读，xlwt写
    @classmethod
    def save_xls_file(cls, fpath, fname, body, head=None, body_keys=None, sflag=0, sheet_name='Sheet1'):
        if not body or not isinstance(body, Iterable):
            return False
        head = cls.get_default(head, 'h')
        try:
            full_path = cls.full_path(fpath, fname)
            book  = xlwt.Workbook(encoding='utf-8', style_compression=0)
            sheet = book.add_sheet(sheet_name, cell_overwrite_ok=True)

            # add simple style
            style = cls.add_style(sheet, head, sflag=sflag)

            # save head
            for j, h_value in enumerate(head):
                sheet.write(0, j, h_value, style)

            # save body
            if isinstance(body[0], dict):
                body_keys = cls.get_default(body_keys, 'b')
                for i, item in enumerate(body):
                    for j, key in enumerate(body_keys):
                        b_value = cls.to_unicode(item.get(key))
                        sheet.write(i+1, j, b_value, style)

            elif isinstance(body[0], (list, tuple)):
                for i in range(0, len(body)):
                    for j in range(0, len(body[i])):
                        b_value = cls.to_unicode(body[i][j])
                        sheet.write(i+1, j, b_value)
            book.save(full_path)
            return True            
        except Exception as e:
            print(e)
            return False

    # 2007版以后的Excel（xlsx结尾的），需要使用openpyxl来读写
    @classmethod
    def save_xlsx07_file(cls, fpath, fname, body, head=None, body_keys=None, sflag='2', sheet_name='Sheet1'):
        if not body or not isinstance(body, Iterable):
            return False
        head = cls.get_default(head, 'h')
        try:
            full_path = cls.full_path(fpath, fname)
            book = openpyxl.Workbook()
            # book = openpyxl.load_workbook(full_path)
            sheet = book.active
            sheet.title = sheet_name

            # add simple style
            style = cls.add_style(sheet, head, sflag='2')

            # save head
            for j, h_value in enumerate(head):
                sheet.cell(row=1, column=j+1, value=h_value)

            # save body
            if isinstance(body[0], dict):
                body_keys = cls.get_default(body_keys, 'b')
                for i, item in enumerate(body):
                    for j, key in enumerate(body_keys):
                        b_value = item.get(key)
                        sheet.cell(row=i+2, column=j+1, value=b_value)
            elif isinstance(body[0], (list, tuple)):
                for i in range(0, len(body)):
                    for j in range(0, len(body[i])):
                        b_value = body[i][j]
                        sheet.cell(row=i+2, column=j+1, value=b_value)
            book.save(full_path)
            return True
        except Exception as e:
            print(e)
            return False


def test_save_xls_file(fpath, fname, body, head):
    ret = ExcelHandle.save_xls_file(fpath, fname, body, head, sflag=1)
    if not ret:
        print("save xls file[{}] failed!".format(fname))
    else:
        print("save xls file[{}] success!".format(fname))


def test_save_xlsx07_file(fpath, fname, body, head):
    ret = ExcelHandle.save_xlsx07_file(fpath, fname, body, head)
    if not ret:
        print("save xls file[{}] failed!".format(fname))
    else:
        print("save xls file[{}] success!".format(fname))


def main():
    f_path = './'
    f_name = 'demo2.xls'
    head = (u'序号', u'地址', u'资产名称', u'添加/发现时间', u'管理员', u'安全状态', u'可用性', u'备注')
    body = [{'id': 1, 'address':'127.0.0.1', 'name':'bbb', 'found_time':'2012/03/06', 'administrator':'admin', 'safe_state':'safe1111', 'description':1}, 
            {'id': 2, 'address':'xxx.x.x.x', 'name':'cc', 'found_time':'2012/03/06', 'administrator':'safe1111', 'safe_state':'safe22', 'description':'你好', 'aa':'nihao'}]

    # test_save_xls_file(f_path, f_name, body, head)
    test_save_xlsx07_file(f_path, f_name, body, head)


if __name__ == "__main__":
    main()
